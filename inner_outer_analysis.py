"""
Analysis: Inner = Outer (non-Generic).
For each Inner GTIN (non-Generic/Placeholder), check if it matches exactly an Outer GTIN.
Split: Same Legal Entity vs Different Legal Entities.
Export: full Inner/Outer rows matched as pairs, for Excel with alternating green highlight.
"""

import polars as pl
from typing import Tuple, List
from io import BytesIO

GENERIC_PLACEHOLDER_GTINS = frozenset({
    "", "9999999999999", "99999999999999", "3000000000009", "30000000000009",
})

COL_LEGAL = "legal entity"
COL_GTIN_INNER = "gtin inner"
COL_GTIN_OUTER = "gtin outer"


def _normalize_gtin_col(ser: pl.Series) -> pl.Series:
    return ser.cast(pl.Utf8).str.replace_all(r"[^0-9]", "")


def build_inner_outer_non_generic(
    df: pl.DataFrame,
    col_legal_entity: str = COL_LEGAL,
    col_gtin_inner: str = COL_GTIN_INNER,
    col_gtin_outer: str = COL_GTIN_OUTER,
) -> Tuple[pl.DataFrame, pl.DataFrame]:
    """
    Returns (same_legal_entity_df, different_legal_entities_df) with columns
    GTIN, Legal Entity (Inner), Legal Entity (Outer).
    """
    df = df.with_columns([
        _normalize_gtin_col(pl.col(col_gtin_inner)).alias("inner_norm"),
        _normalize_gtin_col(pl.col(col_gtin_outer)).alias("outer_norm"),
    ])
    df = df.with_columns([
        pl.when(pl.col("inner_norm").str.len_chars() == 0).then(None).otherwise(pl.col("inner_norm")).alias("inner_norm"),
        pl.when(pl.col("outer_norm").str.len_chars() == 0).then(None).otherwise(pl.col("outer_norm")).alias("outer_norm"),
    ])
    exclude_gtins = list(GENERIC_PLACEHOLDER_GTINS)
    df = df.filter(
        pl.col("inner_norm").is_not_null()
        & pl.col("outer_norm").is_not_null()
        & ~pl.col("inner_norm").is_in(exclude_gtins)
        & ~pl.col("outer_norm").is_in(exclude_gtins)
        & (pl.col("inner_norm").str.replace_all("9", "").str.len_chars() > 0)
        & (pl.col("outer_norm").str.replace_all("9", "").str.len_chars() > 0)
    )
    if len(df) == 0:
        return pl.DataFrame(), pl.DataFrame()

    inner_rows = df.select([
        pl.col("inner_norm").alias("gtin_norm"),
        pl.col(col_legal_entity).alias("le_inner"),
    ]).unique()
    outer_rows = df.select([
        pl.col("outer_norm").alias("gtin_norm"),
        pl.col(col_legal_entity).alias("le_outer"),
    ]).unique()
    matches = inner_rows.join(outer_rows, on="gtin_norm", how="inner")
    if len(matches) == 0:
        return pl.DataFrame(), pl.DataFrame()

    entities_per_gtin = matches.group_by("gtin_norm").agg(
        pl.col("le_inner").implode().alias("le_inner_list"),
        pl.col("le_outer").implode().alias("le_outer_list"),
    ).with_columns(
        pl.col("le_inner_list").list.concat(pl.col("le_outer_list")).list.unique().list.len().alias("num_les")
    )
    same_le_gtins = entities_per_gtin.filter(pl.col("num_les") == 1).select("gtin_norm")
    diff_le_gtins = entities_per_gtin.filter(pl.col("num_les") >= 2).select("gtin_norm")
    same_le_df = matches.filter(pl.col("gtin_norm").is_in(same_le_gtins["gtin_norm"])).unique()
    diff_le_df = matches.filter(pl.col("gtin_norm").is_in(diff_le_gtins["gtin_norm"])).unique()
    same_le_df = same_le_df.rename({"gtin_norm": "GTIN", "le_inner": "Legal Entity (Inner)", "le_outer": "Legal Entity (Outer)"})
    diff_le_df = diff_le_df.rename({"gtin_norm": "GTIN", "le_inner": "Legal Entity (Inner)", "le_outer": "Legal Entity (Outer)"})
    return same_le_df, diff_le_df


def build_inner_outer_export_rows(
    df_stibo: pl.DataFrame,
    col_legal_entity: str = COL_LEGAL,
    col_gtin_inner: str = COL_GTIN_INNER,
    col_gtin_outer: str = COL_GTIN_OUTER,
) -> pl.DataFrame:
    """
    Builds a table of full STIBO rows for each Inner=Outer match, with Role (Inner/Outer) and PairId.
    Each pair is (one row where GTIN is Inner, one row where GTIN is Outer). Rows are ordered
    so that pair 0 = row Inner, row Outer; pair 1 = row Inner, row Outer; etc. for alternating Excel styling.
    """
    df = df_stibo.with_columns([
        _normalize_gtin_col(pl.col(col_gtin_inner)).alias("inner_norm"),
        _normalize_gtin_col(pl.col(col_gtin_outer)).alias("outer_norm"),
    ])
    df = df.with_columns([
        pl.when(pl.col("inner_norm").str.len_chars() == 0).then(None).otherwise(pl.col("inner_norm")).alias("inner_norm"),
        pl.when(pl.col("outer_norm").str.len_chars() == 0).then(None).otherwise(pl.col("outer_norm")).alias("outer_norm"),
    ])
    exclude_gtins = list(GENERIC_PLACEHOLDER_GTINS)
    df = df.filter(
        pl.col("inner_norm").is_not_null()
        & pl.col("outer_norm").is_not_null()
        & ~pl.col("inner_norm").is_in(exclude_gtins)
        & ~pl.col("outer_norm").is_in(exclude_gtins)
        & (pl.col("inner_norm").str.replace_all("9", "").str.len_chars() > 0)
        & (pl.col("outer_norm").str.replace_all("9", "").str.len_chars() > 0)
    )
    if len(df) == 0:
        return pl.DataFrame()

    inner_rows = df.select([pl.col("inner_norm").alias("gtin_norm"), pl.col(col_legal_entity).alias("le_inner")]).unique()
    outer_rows = df.select([pl.col("outer_norm").alias("gtin_norm"), pl.col(col_legal_entity).alias("le_outer")]).unique()
    matches = inner_rows.join(outer_rows, on="gtin_norm", how="inner")
    if len(matches) == 0:
        return pl.DataFrame()

    cols_source = [c for c in df_stibo.columns if c in df.columns and c not in ("inner_norm", "outer_norm")]
    
    # Add pair_id to matches for ordering
    matches = matches.with_row_index("pair_id")
    
    # Prepare inner rows: match on inner_norm = gtin_norm AND legal_entity = le_inner
    inner_df = df.select([
        pl.col("inner_norm").alias("gtin_norm"),
        pl.col(col_legal_entity).alias("le_inner"),
        *[pl.col(c) for c in cols_source]
    ]).filter(pl.col("gtin_norm").is_not_null())
    
    # Prepare outer rows: match on outer_norm = gtin_norm AND legal_entity = le_outer
    outer_df = df.select([
        pl.col("outer_norm").alias("gtin_norm"),
        pl.col(col_legal_entity).alias("le_outer"),
        *[pl.col(c) for c in cols_source]
    ]).filter(pl.col("gtin_norm").is_not_null())
    
    # Join matches with inner rows (on gtin_norm and le_inner)
    inner_joined = matches.join(
        inner_df,
        on=["gtin_norm", "le_inner"],
        how="inner"
    ).select([
        pl.lit("Inner").alias("Role"),
        pl.col("pair_id").alias("Pair"),
        *[pl.col(c) for c in cols_source]
    ])
    
    # Join matches with outer rows (on gtin_norm and le_outer)
    outer_joined = matches.join(
        outer_df,
        on=["gtin_norm", "le_outer"],
        how="inner"
    ).select([
        pl.lit("Outer").alias("Role"),
        pl.col("pair_id").alias("Pair"),
        *[pl.col(c) for c in cols_source]
    ])
    
    # Concatenate and sort by pair_id, then Role (Inner before Outer)
    if len(inner_joined) == 0 and len(outer_joined) == 0:
        return pl.DataFrame()
    
    result = pl.concat([inner_joined, outer_joined]).sort(["Pair", "Role"])
    return result


def export_inner_outer_to_excel_bytes(export_df: pl.DataFrame) -> bytes:
    """
    Writes export_df to Excel and applies alternating green fill per pair:
    pairs 0, 2, 4... green; pairs 1, 3, 5... no fill. Two rows per pair (Inner, Outer).
    Header is row 1, data starts at row 2.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    buffer = BytesIO()
    export_df.to_pandas().to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    max_row = ws.max_row
    max_col = ws.max_column
    for i in range(2, max_row + 1):
        pair_idx = (i - 2) // 2
        if pair_idx % 2 == 0:
            for col in range(1, max_col + 1):
                ws.cell(row=i, column=col).fill = green_fill
    buffer_out = BytesIO()
    wb.save(buffer_out)
    buffer_out.seek(0)
    return buffer_out.getvalue()
